import openpyxl
import datetime

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("rawdata.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

timeline_name_and_rank_csv = open("timelineNameAndRank.csv", "w")

current_line_num = 0

# Iterate the loop to read the cell values
for row in dataframe1.values:
    if current_line_num == 0:

        row_as_list = []

        for value in row:

            if isinstance(value, long):
                row_as_list.append(str(value))
            elif isinstance(value, datetime.date):
                row_as_list.append(str(value))
            elif isinstance(value, bool):
                row_as_list.append(str(value))
            elif isinstance(value, float):
                row_as_list.append(str(value))
            elif value == None:
                row_as_list.append(str(value))
            else:
                row_as_list.append(value.encode("utf-8"))

        index = 0
        for value in row_as_list:
            print(str(index) + ": " + value)
            index += 1

        timeline_name_and_rank_csv.write("Response ID, Timeline Name, Rank\n")
        current_line_num += 1
    else:
        row_as_list = []
        for value in row:

            if isinstance(value, long):
                row_as_list.append(str(value))
            elif isinstance(value, datetime.date):
                row_as_list.append(str(value))
            elif isinstance(value, bool):
                row_as_list.append(str(value))
            elif isinstance(value, float):
                row_as_list.append(str(value))
            elif value == None:
                row_as_list.append(str(value))
            else:
                row_as_list.append(value.encode("utf-8"))

        response_id = row_as_list[0]
        print(response_id)

        timeline1_name = "Timeline 1 (Base Interface)"
        timeline1_rank = row_as_list[126]
        print(timeline1_rank)

        timeline_name_and_rank_csv.write(
            response_id + ", " + timeline1_name + ", " + timeline1_rank + "\n"
        )

        timeline2_name = "Timeline 2 (Density Graph)"
        timeline2_rank = row_as_list[127]
        print(timeline2_rank)

        timeline_name_and_rank_csv.write(
            response_id + ", " + timeline2_name + ", " + timeline2_rank + "\n"
        )

        timeline3_name = "Timeline 3 (Event Blocks)"
        timeline3_rank = row_as_list[128]
        print(timeline3_rank)

        timeline_name_and_rank_csv.write(
            response_id + ", " + timeline3_name + ", " + timeline3_rank + "\n"
        )

        timeline4_name = "Timeline 4 (Density Graph + Event Blocks)"
        timeline4_rank = row_as_list[129]
        print(timeline4_rank)

        timeline_name_and_rank_csv.write(
            response_id + ", " + timeline4_name + ", " + timeline4_rank + "\n"
        )

        timeline5_name = "Timeline 5 (Event Thumbnails)"
        timeline5_rank = row_as_list[130]
        print(timeline5_rank)

        timeline_name_and_rank_csv.write(
            response_id + ", " + timeline5_name + ", " + timeline5_rank + "\n"
        )

# for row in range(0, dataframe1.max_row):
#     row_as_list = [col[row].value for col in dataframe1.iter_cols(1, dataframe1.max_column)]

#     if row == 0:
#         timeline_name_and_rank_csv.write("Timeline Name, Rank\n")
#     else:
#         print(row_as_list)


# timeline_name_and_rank_csv = open("timelineNameAndRank.csv", "w")

# current_line_num = 0

# for timeline_ranks_line in raw_data_lines:
#     timeline_ranks_line = timeline_ranks_line.strip().split(",")
#     print(timeline_ranks_line)
#     if current_line_num == 0:

#         timeline_name_and_rank_csv.write("Timeline Name, Rank\n")

#         current_line_num += 1
#     else:

#         timeline1_name = "Timeline 1 (Base Interface)"
#         timeline1_rank = timeline_ranks_line[0]
#         print(timeline1_rank)

#         timeline_name_and_rank_csv.write(timeline1_name + ", " + timeline1_rank + "\n")

#         timeline2_name = "Timeline 2 (Density Graph)"
#         timeline2_rank = timeline_ranks_line[1]
#         print(timeline2_rank)

#         timeline_name_and_rank_csv.write(timeline2_name + ", " + timeline2_rank + "\n")

#         timeline3_name = "Timeline 3 (Event Blocks)"
#         timeline3_rank = timeline_ranks_line[2]
#         print(timeline3_rank)

#         timeline_name_and_rank_csv.write(timeline3_name + ", " + timeline3_rank + "\n")

#         timeline4_name = "Timeline 4 (Density Graph + Event Blocks)"
#         timeline4_rank = timeline_ranks_line[3]
#         print(timeline4_rank)

#         timeline_name_and_rank_csv.write(timeline4_name + ", " + timeline4_rank + "\n")

#         timeline5_name = "Timeline 5 (Event Thumbnails)"
#         timeline5_rank = timeline_ranks_line[4]
#         print(timeline5_rank)

#         timeline_name_and_rank_csv.write(timeline5_name + ", " + timeline5_rank + "\n")

# raw_data_lines.close()
timeline_name_and_rank_csv.close()
